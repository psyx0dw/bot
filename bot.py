# coding: utf-8
import telebot
from telebot import types
import sqlite3
import threading
import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import time
# --- –ö–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏—è (–ø—Ä–µ–¥–ø–æ–ª–∞–≥–∞–µ—Ç—Å—è, —á—Ç–æ —ç—Ç–∏ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω—ã –≤ config.py) ---
from config import BOT_TOKEN, ADMIN_GROUP_ID, BONUS_PERCENT, MAX_DISCOUNT, REFERRAL_BONUS, DB_PATH
bot = telebot.TeleBot(BOT_TOKEN)
# –õ–æ–∫–∞–ª—å–Ω–æ–µ —Ö—Ä–∞–Ω–∏–ª–∏—â–µ –¥–ª—è –ø–æ—Ç–æ–∫–æ–≤ –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è–º–∏ —Å –ë–î
local_storage = threading.local()
class DBManager:
    """–ö–ª–∞—Å—Å –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –≤—Å–µ–º–∏ –æ–ø–µ—Ä–∞—Ü–∏—è–º–∏ —Å –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö."""
    def __init__(self, db_path):
        self.db_path = db_path
        self._init_db_if_not_exists()
    def get_conn(self):
        """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–µ —Å –±–∞–∑–æ–π –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —Ç–µ–∫—É—â–µ–≥–æ –ø–æ—Ç–æ–∫–∞."""
        if not hasattr(local_storage, 'conn'):
            local_storage.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        return local_storage.conn
    def _init_db_if_not_exists(self, sql_file="models.sql"):
        """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ—Ç –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö –∏–∑ SQL-—Å–∫—Ä–∏–ø—Ç–∞, –µ—Å–ª–∏ —Ñ–∞–π–ª –ë–î –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç."""
        try:
            if not os.path.exists(self.db_path):
                with open(sql_file, 'r', encoding='utf-8') as f:
                    sql_script = f.read()
               
                with self.get_conn() as conn:
                    conn.executescript(sql_script)
                print(f"‚úÖ Database tables created from {sql_file}.")
            else:
                print("‚úÖ Database file already exists.")
        except FileNotFoundError:
            print(f"‚ùå Error: Database schema file '{sql_file}' not found.")
            print("Please ensure models.sql is in the same directory.")
        except Exception as e:
            print(f"‚ùå An error occurred during database initialization: {e}")
    def get_user(self, telegram_id):
        """
        –ü–æ–ª—É—á–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è, –≤–∫–ª—é—á–∞—è –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ä–µ—Ñ–µ—Ä–∞–ª–æ–≤ –∏ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–∫–∞–∑–æ–≤.
        –û–±–Ω–æ–≤–ª–µ–Ω–æ –¥–ª—è –≤—ã–≤–æ–¥–∞ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –∑–∞–∫–∞–∑–æ–≤ –≤ –ø—Ä–æ—Ñ–∏–ª–µ.
        """
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT u.*,
                       (SELECT COUNT(*) FROM users WHERE referrer_id = u.id) AS referrals,
                       (SELECT COUNT(*) FROM orders WHERE user_id = u.id) AS order_count
                FROM users u WHERE telegram_id = ?
            """, (telegram_id,))
            return cur.fetchone()
    def get_user_id(self, telegram_id):
        """–ü–æ–ª—É—á–∞–µ—Ç ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ –µ–≥–æ Telegram ID."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM users WHERE telegram_id = ?", (telegram_id,))
            row = cur.fetchone()
            return row[0] if row else None
    def add_user_full(self, telegram_id, name, referrer_telegram_id=None):
        """–î–æ–±–∞–≤–ª—è–µ—Ç –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            referrer_id = None
            if referrer_telegram_id:
                cur.execute("SELECT id FROM users WHERE telegram_id = ?", (referrer_telegram_id,))
                r = cur.fetchone()
                if r:
                    referrer_id = r[0]
            cur.execute("""
                INSERT OR IGNORE INTO users (telegram_id, name, referrer_id)
                VALUES (?, ?, ?)
            """, (telegram_id, name, referrer_id))
            conn.commit()
    def update_points(self, telegram_id, delta):
        """–û–±–Ω–æ–≤–ª—è–µ—Ç –±–∞–ª–ª—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è."""
        with self.get_conn() as conn:
            conn.execute("UPDATE users SET points = COALESCE(points, 0) + ? WHERE telegram_id = ?", (delta, telegram_id))
            conn.commit()
    def get_menu(self):
        """–ü–æ–ª—É—á–∞–µ—Ç –≤—Å—ë –º–µ–Ω—é."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT category, name, size, price, quantity FROM stock ORDER BY category, name, CAST(size AS REAL)")
            return cur.fetchall()
    def get_stock_by_fullname(self, full_name):
        """–ü–æ–ª—É—á–∞–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–∞ –ø–æ –µ–≥–æ –ø–æ–ª–Ω–æ–º—É –Ω–∞–∑–≤–∞–Ω–∏—é."""
        if not full_name.endswith("–ª"):
            return 0
        name, size_l = full_name.rsplit(" ", 1)
        size = size_l.replace("–ª", "")
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT price, quantity FROM stock WHERE LOWER(name)=LOWER(?) AND size=?", (name.strip(), size))
            row = cur.fetchone()
            return row[0] if row else None, row[1] if row else 0
    def get_stock_by_name_size(self, name, size):
        """–ü–æ–ª—É—á–∞–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —Ç–æ–≤–∞—Ä–µ –ø–æ –∏–º–µ–Ω–∏ –∏ —Ä–∞–∑–º–µ—Ä—É."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT price, quantity FROM stock WHERE LOWER(name)=LOWER(?) AND size=?", (name.strip(), size))
            return cur.fetchone()
    def reduce_stock(self, full_name, qty):
        """–£–º–µ–Ω—å—à–∞–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–∞ –Ω–∞ —Å–∫–ª–∞–¥–µ."""
        if not full_name.endswith("–ª"):
            return
        name, size_l = full_name.rsplit(" ", 1)
        size = size_l.replace("–ª", "")
        with self.get_conn() as conn:
            conn.execute("""
                UPDATE stock SET quantity = quantity - ?
                WHERE name = ? AND size = ? AND quantity >= ?
            """, (qty, name, size, qty))
            conn.commit()
            cur = conn.cursor()
            cur.execute("SELECT quantity FROM stock WHERE name=? AND size=?", (name, size))
            row = cur.fetchone()
            if row and row[0] < 3:
                bot.send_message(ADMIN_GROUP_ID, f"‚ö†Ô∏è –û—Å—Ç–∞—Ç–æ–∫ –Ω–∏–∑–∫–∏–π: <b>{name} {size}–ª</b> ‚Äî {row[0]} —à—Ç", parse_mode="HTML")
    def get_cart_items(self, telegram_id):
        """–ü–æ–ª—É—á–∞–µ—Ç –≤—Å–µ –ø–æ–∑–∏—Ü–∏–∏ –∏–∑ –∫–æ—Ä–∑–∏–Ω—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT name, price, qty
                FROM cart
                WHERE user_id = (SELECT id FROM users WHERE telegram_id = ?)
            """, (telegram_id,))
            return [{"name": r[0], "price": r[1], "qty": r[2]} for r in cur.fetchall()]
    def add_to_cart(self, telegram_id, item_name, price, qty):
        """–î–æ–±–∞–≤–ª—è–µ—Ç –ø–æ–∑–∏—Ü–∏—é –≤ –∫–æ—Ä–∑–∏–Ω—É –≤ –ë–î."""
        user_id = self.get_user_id(telegram_id)
        if not user_id: return
        with self.get_conn() as conn:
            conn.execute("""
                INSERT INTO cart (user_id, name, price, qty)
                VALUES (?, ?, ?, ?)
            """, (user_id, item_name, price, qty))
            conn.commit()
    def clear_cart(self, telegram_id):
        """–û—á–∏—â–∞–µ—Ç –∫–æ—Ä–∑–∏–Ω—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –ë–î."""
        with self.get_conn() as conn:
            conn.execute("DELETE FROM cart WHERE user_id = (SELECT id FROM users WHERE telegram_id = ?)", (telegram_id,))
            conn.commit()
    def create_order(self, telegram_id, items, total):
        """–°–æ–∑–¥–∞–µ—Ç –Ω–æ–≤—ã–π –∑–∞–∫–∞–∑ –≤ –ë–î."""
        user_id = self.get_user_id(telegram_id)
        if not user_id: return None
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("INSERT INTO orders (user_id, total, created_at) VALUES (?, ?, ?)", (user_id, total, datetime.datetime.now()))
            order_id = cur.lastrowid
            for it in items:
                cur.execute(
                    "INSERT INTO order_items (order_id, name, price, qty) VALUES (?, ?, ?, ?)",
                    (order_id, it["name"], it["price"], it["qty"])
                )
            conn.commit()
            return order_id
   
    def get_admin_data(self):
        """–ü–æ–ª—É—á–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –¥–ª—è –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*), SUM(total) FROM orders")
            count, revenue = cur.fetchone()
            cur.execute("SELECT COUNT(*) FROM users")
            users = cur.fetchone()[0]
            return count, revenue, users
    def get_low_stock(self):
        """–ü–æ–ª—É—á–∞–µ—Ç —Ç–æ–≤–∞—Ä—ã —Å –Ω–∏–∑–∫–∏–º –æ—Å—Ç–∞—Ç–∫–æ–º."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT category, name, size, quantity FROM stock WHERE quantity < 3 ORDER BY quantity")
            return cur.fetchall()
    def get_recent_orders(self):
        """–ü–æ–ª—É—á–∞–µ—Ç —Å–ø–∏—Å–æ–∫ –ø–æ—Å–ª–µ–¥–Ω–∏—Ö 10 –∑–∞–∫–∞–∑–æ–≤."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT o.id, u.name, o.total, o.created_at
                FROM orders o JOIN users u ON o.user_id = u.id
                ORDER BY o.created_at DESC LIMIT 10
            """)
            return cur.fetchall()
    def admin_update_item(self, cat, name, size, price, qty):
        """–î–æ–±–∞–≤–ª—è–µ—Ç/–æ–±–Ω–æ–≤–ª—è–µ—Ç —Ç–æ–≤–∞—Ä –≤—Ä—É—á–Ω—É—é."""
        with self.get_conn() as conn:
            conn.execute("INSERT OR REPLACE INTO stock (category, name, size, price, quantity) VALUES (?, ?, ?, ?, ?)",
                         (cat, name, size, price, qty))
            conn.commit()
    def admin_update_qty(self, name, size, new_qty):
        """–û–±–Ω–æ–≤–ª—è–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–∞."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("UPDATE stock SET quantity = ? WHERE name = ? AND size = ?", (new_qty, name, size))
            conn.commit()
            return cur.rowcount > 0
    def admin_delete_item(self, name, size):
        """–£–¥–∞–ª—è–µ—Ç —Ç–æ–≤–∞—Ä."""
        with self.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM stock WHERE name = ? AND size = ?", (name, size))
            conn.commit()
            return cur.rowcount > 0
db_manager = DBManager(DB_PATH)
def calc_discount(total, points):
    """–†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ—Ç —Å–∫–∏–¥–∫—É –Ω–∞ –æ—Å–Ω–æ–≤–µ –±–∞–ª–ª–æ–≤."""
    return min(points, int(total * MAX_DISCOUNT))
def format_cart_lines(items):
    """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç —Å–ø–∏—Å–æ–∫ –ø–æ–∑–∏—Ü–∏–π –≤ –∫–æ—Ä–∑–∏–Ω–µ."""
    return "\n".join([f"‚Ä¢ {it['name']} x{it['qty']} ‚Äî {it['price']}‚ÇΩ" for it in items])
# --- –ö–ª–∞–≤–∏–∞—Ç—É—Ä—ã ---
def main_keyboard():
    """–û—Å–Ω–æ–≤–Ω–∞—è –∫–ª–∞–≤–∏–∞—Ç—É—Ä–∞ –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞."""
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("üìã –ú–µ–Ω—é", "üõí –ö–æ—Ä–∑–∏–Ω–∞")
    kb.add("üë§ –ú–æ–π –ø—Ä–æ—Ñ–∏–ª—å", "üîó –†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –ø—Ä–æ–≥—Ä–∞–º–º–∞")
    kb.add("üõ† –¢–µ—Ö–ø–æ–¥–¥–µ—Ä–∂–∫–∞")
    return kb
def handle_add_more_keyboard():
    """–ö–ª–∞–≤–∏–∞—Ç—É—Ä–∞ –¥–ª—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –Ω–æ–≤—ã—Ö –ø–æ–∑–∏—Ü–∏–π –≤ –∫–æ—Ä–∑–∏–Ω—É."""
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("‚ûï –î–æ–±–∞–≤–∏—Ç—å –µ—â—ë", "‚úÖ –ü–µ—Ä–µ–π—Ç–∏ –∫ –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—é")
    return kb
# --- –û–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –∫–æ–º–∞–Ω–¥ –∏ —Å–æ–æ–±—â–µ–Ω–∏–π ---
@bot.message_handler(commands=["start"])
def start(message):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /start."""
    telegram_id = str(message.chat.id)
    referrer_telegram_id = None
    if message.text.startswith("/start "):
        try:
            referrer_telegram_id = message.text.split(" ", 1)[1]
            if referrer_telegram_id == telegram_id or not db_manager.get_user(referrer_telegram_id):
                bot.send_message(telegram_id, "üòÖ –ü–æ—Ö–æ–∂–µ, —Å—Å—ã–ª–∫–∞ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–∞ –∏–ª–∏ —ç—Ç–æ —Ç–≤–æ–π —Å–æ–±—Å—Ç–≤–µ–Ω–Ω—ã–π ID. –†–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—è –±–µ–∑ —Ä–µ—Ñ–µ—Ä–∞–ª–∞.", parse_mode="HTML")
                referrer_telegram_id = None
            else:
                bot.send_message(telegram_id, "ü§ù –û—Ç–ª–∏—á–Ω–æ ‚Äî —Ç—ã –ø—Ä–∏—à—ë–ª –ø–æ —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω–æ–π —Å—Å—ã–ª–∫–µ! –î–∞–≤–∞–π –ø–æ–∑–Ω–∞–∫–æ–º–∏–º—Å—è.", parse_mode="HTML")
        except IndexError:
            pass
   
    user = db_manager.get_user(telegram_id)
    if user:
        bot.send_message(telegram_id, f"‚òï –° –≤–æ–∑–≤—Ä–∞—â–µ–Ω–∏–µ–º, <b>{user[2]}</b>!", reply_markup=main_keyboard(), parse_mode="HTML")
    else:
        msg = bot.send_message(telegram_id, "‚òï –ü—Ä–∏–≤–µ—Ç! –Ø –ö–æ—Ñ–µ–π–Ω—ã–π!. –ö–∞–∫ —Ç–µ–±—è –∑–æ–≤—É—Ç?", parse_mode="HTML")
        bot.register_next_step_handler(msg, finish_registration, referrer_telegram_id)
def finish_registration(message, referrer_telegram_id=None):
    """–ó–∞–≤–µ—Ä—à–∞–µ—Ç —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—é –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è."""
    name = message.text.strip()
    if len(name) < 2:
        msg = bot.send_message(message.chat.id, "‚ùå –ò–º—è —Å–ª–∏—à–∫–æ–º –∫–æ—Ä–æ—Ç–∫–æ–µ. –ù–∞–ø–∏—à–∏, –ø–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–ª–Ω–æ–µ –∏–º—è.", parse_mode="HTML")
        bot.register_next_step_handler(msg, finish_registration, referrer_telegram_id)
        return
   
    telegram_id = str(message.chat.id)
    db_manager.add_user_full(telegram_id, name, referrer_telegram_id=referrer_telegram_id)
   
    bot.send_message(
        telegram_id,
        f"üéâ –û—Ç–ª–∏—á–Ω–æ, <b>{name}</b>! –î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å –≤ –Ω–∞—à—É –∫–æ—Ñ–µ–π–Ω—é ‚òï\n\n"
        "–¢–µ–ø–µ—Ä—å —Ç—ã –º–æ–∂–µ—à—å –ø—Ä–æ—Å–º–∞—Ç—Ä–∏–≤–∞—Ç—å –º–µ–Ω—é, —Å–æ–±–∏—Ä–∞—Ç—å –∫–æ—Ä–∑–∏–Ω—É –∏ –æ—Ñ–æ—Ä–º–ª—è—Ç—å –∑–∞–∫–∞–∑—ã –ø—Ä—è–º–æ –∑–¥–µ—Å—å.",
        reply_markup=main_keyboard(),
        parse_mode="HTML"
    )
    bot.send_message(
        telegram_id,
        "üìå <b>–ö–∞–∫ —Å–¥–µ–ª–∞—Ç—å –∑–∞–∫–∞–∑:</b>\n\n"
        "1Ô∏è‚É£ –ù–∞–∂–º–∏ ¬´üìã –ú–µ–Ω—é¬ª –∏ –≤—ã–±–µ—Ä–∏ –Ω–∞–ø–∏—Ç–æ–∫\n"
        "2Ô∏è‚É£ –ù–∞–ø–∏—à–∏ –Ω–∞–∑–≤–∞–Ω–∏–µ –∏ –æ–±—ä—ë–º, –Ω–∞–ø—Ä–∏–º–µ—Ä: <i>–õ–∞—Ç—Ç–µ 0.3–ª</i>\n"
        "3Ô∏è‚É£ –ü–µ—Ä–µ–π–¥–∏ –≤ ¬´üõí –ö–æ—Ä–∑–∏–Ω–∞¬ª –∏ –ø–æ–¥—Ç–≤–µ—Ä–¥–∏ –∑–∞–∫–∞–∑\n\n"
        "–û–±—ã—á–Ω–æ –ø—Ä–∏–≥–æ—Ç–æ–≤–ª–µ–Ω–∏–µ –∑–∞–Ω–∏–º–∞–µ—Ç 5‚Äì10 –º–∏–Ω—É—Ç.",
        parse_mode="HTML"
    )
@bot.message_handler(func=lambda m: m.text == "üìã –ú–µ–Ω—é")
def show_menu(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –∫–ª–∏–µ–Ω—Ç—É —Ç–µ–∫—É—â–µ–µ –º–µ–Ω—é."""
    rows = db_manager.get_menu()
    if not rows:
        bot.send_message(message.chat.id, "üì≠ –ú–µ–Ω—é –ø–æ–∫–∞ –ø—É—Å—Ç–æ–µ.", parse_mode="HTML")
        return
    text = "üìã <b>–ù–∞—à–µ –º–µ–Ω—é:</b>\n"
    current_cat = None
    for cat, name, size, price, qty in rows:
        if cat != current_cat:
            text += f"\nüî∏ <b>{cat}</b>\n"
            current_cat = cat
        size_str = f" {size}–ª" if size else ""
        text += f"‚Ä¢ {name}{size_str} ‚Äî {price}‚ÇΩ (–æ—Å—Ç–∞–ª–æ—Å—å: {qty})\n"
    bot.send_message(message.chat.id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üõí –ö–æ—Ä–∑–∏–Ω–∞")
def show_cart(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ –∫–æ—Ä–∑–∏–Ω—ã –∫–ª–∏–µ–Ω—Ç–∞."""
    telegram_id = str(message.chat.id)
    user = db_manager.get_user(telegram_id)
    if not user:
        bot.send_message(telegram_id, "‚ö†Ô∏è –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, —Å–Ω–∞—á–∞–ª–∞ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–π—Ç–µ—Å—å ‚Äî –Ω–∞–∂–º–∏ /start.", parse_mode="HTML")
        return
    items = db_manager.get_cart_items(telegram_id)
    if not items:
        bot.send_message(telegram_id, "üõí –¢–≤–æ—è –∫–æ—Ä–∑–∏–Ω–∞ –ø—É—Å—Ç–∞. –î–æ–±–∞–≤—å –Ω–∞–ø–∏—Ç–∫–∏ –∏–∑ –º–µ–Ω—é.", parse_mode="HTML")
        return
   
    total = sum(it["price"] * it["qty"] for it in items)
    points = user[5] or 0
    referrals = user[-2] or 0 # –ò–∑–º–µ–Ω–µ–Ω–æ –Ω–∞ -2, —Ç–∞–∫ –∫–∞–∫ –¥–æ–±–∞–≤–∏–ª–æ—Å—å –ø–æ–ª–µ orders_count
   
    is_challenge_eligible = points >= 500 and referrals >= 10
   
    kb = types.InlineKeyboardMarkup()
   
    if is_challenge_eligible:
        kb.add(types.InlineKeyboardButton("üéÅ –ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞—Ç—å 15% —Å–∫–∏–¥–∫—É", callback_data="activate_15_percent"))
   
    standard_discount = calc_discount(total, points)
    final_total_standard = total - standard_discount
   
    kb.add(types.InlineKeyboardButton("‚úÖ –û—Ñ–æ—Ä–º–∏—Ç—å –∑–∞–∫–∞–∑", callback_data="confirm_standard"))
    kb.add(types.InlineKeyboardButton("üßπ –û—á–∏—Å—Ç–∏—Ç—å –∫–æ—Ä–∑–∏–Ω—É", callback_data="clear"))
    bot.send_message(
        telegram_id,
        f"üõí <b>–¢–≤–æ—è –∫–æ—Ä–∑–∏–Ω–∞:</b>\n{format_cart_lines(items)}\n\n"
        f"üí∞ <b>–ò—Ç–æ–≥–æ:</b> {total}‚ÇΩ\n"
        f"üéÅ <b>–°–∫–∏–¥–∫–∞ –ø–æ –±–∞–ª–ª–∞–º:</b> {standard_discount}‚ÇΩ\n"
        f"üì¶ <b>–ö –æ–ø–ª–∞—Ç–µ:</b> {final_total_standard}‚ÇΩ",
        reply_markup=kb,
        parse_mode="HTML"
    )
@bot.callback_query_handler(func=lambda call: call.data == "activate_15_percent")
def activate_15_percent(call):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –∞–∫—Ç–∏–≤–∞—Ü–∏—é 15% —Å–∫–∏–¥–∫–∏."""
    telegram_id = str(call.message.chat.id)
    user = db_manager.get_user(telegram_id)
    items = db_manager.get_cart_items(telegram_id)
   
    if not user or not items:
        bot.answer_callback_query(call.id, "–ö–æ—Ä–∑–∏–Ω–∞ –ø—É—Å—Ç–∞ –∏–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return
       
    points = user[5] or 0
    referrals = user[-2] or 0 # –ò–∑–º–µ–Ω–µ–Ω–æ –Ω–∞ -2, —Ç–∞–∫ –∫–∞–∫ –¥–æ–±–∞–≤–∏–ª–æ—Å—å –ø–æ–ª–µ orders_count
    if not (points >= 500 and referrals >= 10):
        bot.answer_callback_query(call.id, "‚ùå –£—Å–ª–æ–≤–∏—è –¥–ª—è 15% —Å–∫–∏–¥–∫–∏ –Ω–µ –≤—ã–ø–æ–ª–Ω–µ–Ω—ã.")
        return
    total = sum(it["price"] * it["qty"] for it in items)
    discount = int(total * 0.15)
    final_total = total - discount
   
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("‚úÖ –ü–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –∑–∞–∫–∞–∑", callback_data="confirm_challenge"))
    kb.add(types.InlineKeyboardButton("‚ùå –û—Ç–º–µ–Ω–∞", callback_data="clear"))
   
    bot.edit_message_text(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text=(
            f"üéâ <b>–ß–µ–ª–ª–µ–Ω–¥–∂ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω!</b>\n\n"
            f"üéÅ –¢—ã –ø–æ–ª—É—á–∞–µ—à—å —Å–∫–∏–¥–∫—É <b>15%</b> –Ω–∞ —ç—Ç–æ—Ç –∑–∞–∫–∞–∑.\n"
            f"üí∞ –ò—Ç–æ–≥–æ: {total}‚ÇΩ\n"
            f"üìâ –°–∫–∏–¥–∫–∞: {discount}‚ÇΩ\n"
            f"üì¶ –ö –æ–ø–ª–∞—Ç–µ: {final_total}‚ÇΩ\n\n"
            "‚ö†Ô∏è –î–ª—è —ç—Ç–æ–π –æ–ø–µ—Ä–∞—Ü–∏–∏ –±—É–¥–µ—Ç —Å–ø–∏—Å–∞–Ω–æ <b>500</b> –±–∞–ª–ª–æ–≤."
        ),
        reply_markup=kb,
        parse_mode="HTML"
    )
@bot.callback_query_handler(func=lambda call: call.data == "confirm_standard")
def confirm_order_standard(call):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ –∑–∞–∫–∞–∑–∞ —Å–æ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–π —Å–∫–∏–¥–∫–æ–π."""
    telegram_id = str(call.message.chat.id)
    user = db_manager.get_user(telegram_id)
    items = db_manager.get_cart_items(telegram_id)
   
    if not user or not items:
        bot.answer_callback_query(call.id, "–ö–æ—Ä–∑–∏–Ω–∞ –ø—É—Å—Ç–∞ –∏–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return
   
    points = user[5] or 0
    total = sum(it["price"] * it["qty"] for it in items)
    discount = calc_discount(total, points)
    final_total = total - discount
   
    process_order(telegram_id, user, items, final_total, discount)
    try:
        bot.edit_message_reply_markup(call.message.chat.id, call.message.id, reply_markup=None)
    except Exception:
        pass
@bot.callback_query_handler(func=lambda call: call.data == "confirm_challenge")
def confirm_order_challenge(call):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ –∑–∞–∫–∞–∑–∞ —Å 15% —Å–∫–∏–¥–∫–æ–π (—á–µ–ª–ª–µ–Ω–¥–∂)."""
    telegram_id = str(call.message.chat.id)
    user = db_manager.get_user(telegram_id)
    items = db_manager.get_cart_items(telegram_id)
    if not user or not items:
        bot.answer_callback_query(call.id, "–ö–æ—Ä–∑–∏–Ω–∞ –ø—É—Å—Ç–∞ –∏–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω.")
        return
    points = user[5] or 0
    if points < 500:
        bot.answer_callback_query(call.id, "‚ùå –ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –±–∞–ª–ª–æ–≤ –¥–ª—è –∞–∫—Ç–∏–≤–∞—Ü–∏–∏ —Å–∫–∏–¥–∫–∏.")
        return
       
    total = sum(it["price"] * it["qty"] for it in items)
    discount = int(total * 0.15)
    final_total = total - discount
   
    process_order(telegram_id, user, items, final_total, discount, points_spent=500)
    try:
        bot.edit_message_reply_markup(call.message.chat.id, call.message.id, reply_markup=None)
    except Exception:
        pass
def process_order(telegram_id, user, items, final_total, discount, points_spent=0):
    """–û–±—â–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∏ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è –∑–∞–∫–∞–∑–∞."""
    for it in items:
        _, qty_in_stock = db_manager.get_stock_by_fullname(it["name"])
        if qty_in_stock < it["qty"]:
            bot.send_message(telegram_id, f"‚ùå –ò–∑–≤–∏–Ω–∏—Ç–µ, <b>{it['name']}</b> —Å–µ–π—á–∞—Å –∑–∞–∫–æ–Ω—á–∏–ª—Å—è.", parse_mode="HTML")
            return
           
    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –ø–µ—Ä–≤—ã–π –∑–∞–∫–∞–∑ –¥–ª—è –Ω–∞—á–∏—Å–ª–µ–Ω–∏—è —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω–æ–≥–æ –±–æ–Ω—É—Å–∞
    user_id = db_manager.get_user_id(telegram_id)
    with db_manager.get_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM orders WHERE user_id = ?", (user_id,))
        orders_count = cur.fetchone()[0]
   
    # –ï—Å–ª–∏ —ç—Ç–æ –ø–µ—Ä–≤—ã–π –∑–∞–∫–∞–∑ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è, –∏ —É –Ω–µ–≥–æ –µ—Å—Ç—å —Ä–µ—Ñ–µ—Ä–µ—Ä
    if orders_count == 0:
        with db_manager.get_conn() as conn:
            cur = conn.cursor()
            cur.execute("SELECT referrer_id FROM users WHERE id = ?", (user_id,))
            referrer_row = cur.fetchone()
            referrer_id = referrer_row[0] if referrer_row else None
       
        if referrer_id:
            # –ù–∞—á–∏—Å–ª–µ–Ω–∏–µ –±–æ–Ω—É—Å–æ–≤ –∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
            db_manager.update_points(telegram_id, 50) # –±–æ–Ω—É—Å –Ω–æ–≤–æ–º—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
            with db_manager.get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT telegram_id FROM users WHERE id = ?", (referrer_id,))
                referrer_telegram_id = cur.fetchone()[0]
            db_manager.update_points(referrer_telegram_id, REFERRAL_BONUS)
            bot.send_message(referrer_telegram_id, f"üéâ –û—Ç–ª–∏—á–Ω–æ! –¢–≤–æ–π –¥—Ä—É–≥ <b>{user[2]}</b> —Å–¥–µ–ª–∞–ª –ø–µ—Ä–≤—ã–π –∑–∞–∫–∞–∑ ‚Äî —Ç—ã –ø–æ–ª—É—á–∏–ª {REFERRAL_BONUS} –±–∞–ª–ª–æ–≤!", parse_mode="HTML")
           
            # –£–¥–∞–ª–µ–Ω–∏–µ —Ä–µ—Ñ–µ—Ä–µ—Ä–∞, —á—Ç–æ–±—ã –±–æ–Ω—É—Å –±—ã–ª –µ–¥–∏–Ω–æ—Ä–∞–∑–æ–≤—ã–º
            with db_manager.get_conn() as conn:
                conn.execute("UPDATE users SET referrer_id = NULL WHERE id = ?", (user_id,))
                conn.commit()
    for it in items:
        db_manager.reduce_stock(it["name"], it["qty"])
   
    earned = int(final_total * BONUS_PERCENT)
    db_manager.update_points(telegram_id, earned - points_spent)
    db_manager.clear_cart(telegram_id)
   
    order_id = db_manager.create_order(telegram_id, items, final_total)
   
    items_text = "; ".join([f"{it['name']} x{it['qty']}" for it in items])
    bot.send_message(
        telegram_id,
        f"‚úÖ <b>–ó–∞–∫–∞–∑ ‚Ññ{order_id} –æ—Ñ–æ—Ä–º–ª–µ–Ω!</b>\n\n"
        f"üßæ –ü–æ–∑–∏—Ü–∏–∏: {items_text}\n"
        f"üí≥ –ö –æ–ø–ª–∞—Ç–µ: {final_total}‚ÇΩ\n"
        f"üéØ –ë–∞–ª–ª—ã: +{earned} / -{points_spent}\n\n"
        "‚òï –°–ø–∞—Å–∏–±–æ! –ó–∞–∫–∞–∑ –≥–æ—Ç–æ–≤–∏—Ç—Å—è ‚Äî –∑–∞–π–¥–∏ –Ω–∞ —Ç–æ—á–∫—É —á–µ—Ä–µ–∑ 5‚Äì10 –º–∏–Ω—É—Ç, —á—Ç–æ–±—ã –∑–∞–±—Ä–∞—Ç—å.",
        reply_markup=main_keyboard(),
        parse_mode="HTML"
    )
   
    admin_message_text = (
        f"üì¶ <b>–ù–æ–≤—ã–π –∑–∞–∫–∞–∑ ‚Ññ{order_id}</b>\n"
        f"üë§ –ö–ª–∏–µ–Ω—Ç: {user[2]}\n"
        f"üßæ –ü–æ–∑–∏—Ü–∏–∏: {items_text}\n"
        f"üí∞ –ö –æ–ø–ª–∞—Ç–µ: {final_total}‚ÇΩ\n"
        f"üìâ –°–∫–∏–¥–∫–∞: {discount}‚ÇΩ"
    )
   
    ready_button = types.InlineKeyboardMarkup()
    ready_button.add(types.InlineKeyboardButton("‚úÖ –ì–æ—Ç–æ–≤", callback_data=f"ready_{telegram_id}_{order_id}"))
    ready_button.add(types.InlineKeyboardButton("üí¨ –°–≤—è–∑–∞—Ç—å—Å—è —Å –∫–ª–∏–µ–Ω—Ç–æ–º", callback_data=f"contact_{telegram_id}_{order_id}"))
    bot.send_message(
        ADMIN_GROUP_ID,
        admin_message_text,
        reply_markup=ready_button,
        parse_mode="HTML"
    )
@bot.callback_query_handler(func=lambda call: call.data == "clear")
def clear_cart_handler(call):
    """–û—á–∏—â–∞–µ—Ç –∫–æ—Ä–∑–∏–Ω—É –∫–ª–∏–µ–Ω—Ç–∞."""
    telegram_id = str(call.message.chat.id)
    db_manager.clear_cart(telegram_id)
    bot.answer_callback_query(call.id, "üóë –ö–æ—Ä–∑–∏–Ω–∞ –æ—á–∏—â–µ–Ω–∞.")
    bot.send_message(telegram_id, "üß∫ –í–∞—à–∞ –∫–æ—Ä–∑–∏–Ω–∞ –ø—É—Å—Ç–∞. –ß–µ–º –µ—â—ë –ø–æ–º–æ—á—å?", reply_markup=main_keyboard(), parse_mode="HTML")
@bot.callback_query_handler(func=lambda call: call.data.startswith("ready_"))
def mark_ready(call):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–ª—è –∫–Ω–æ–ø–∫–∏ '–ì–æ—Ç–æ–≤' –≤ –∞–¥–º–∏–Ω-–≥—Ä—É–ø–ø–µ."""
    _, telegram_id, order_id = call.data.split("_", 2)
    bot.send_message(int(telegram_id), f"‚úÖ –í–∞—à –∑–∞–∫–∞–∑ ‚Ññ{order_id} –≥–æ—Ç–æ–≤! –ó–∞–±–µ—Ä–∏—Ç–µ –µ–≥–æ –Ω–∞ —Ç–æ—á–∫–µ ‚òï")
    bot.answer_callback_query(call.id, "–ö–ª–∏–µ–Ω—Ç —É–≤–µ–¥–æ–º–ª—ë–Ω ‚úÖ")
    try:
        bot.edit_message_reply_markup(call.message.chat.id, call.message.id, reply_markup=None)
    except Exception:
        pass
@bot.callback_query_handler(func=lambda call: call.data.startswith("contact_"))
def contact_user_handler(call):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–ª—è –∫–Ω–æ–ø–∫–∏ '–°–≤—è–∑–∞—Ç—å—Å—è —Å –∫–ª–∏–µ–Ω—Ç–æ–º'."""
    try:
        _, telegram_id, order_id = call.data.split("_", 2)
        msg = bot.send_message(call.message.chat.id, f"‚úâÔ∏è –ù–∞–ø–∏—à–∏ —Å–æ–æ–±—â–µ–Ω–∏–µ –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ (–∑–∞–∫–∞–∑ #{order_id}):", parse_mode="HTML")
        bot.register_next_step_handler(msg, send_admin_message, telegram_id, order_id)
    except Exception as e:
        bot.send_message(call.message.chat.id, f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ø—ã—Ç–∫–µ —Å–≤—è–∑–∞—Ç—å—Å—è: {e}", parse_mode="HTML")
def send_admin_message(message, telegram_id, order_id):
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç —Å–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç –∞–¥–º–∏–Ω–∞ –∫–ª–∏–µ–Ω—Ç—É."""
    try:
        user_msg = message.text
        bot.send_message(telegram_id, f"üìù –°–æ–æ–±—â–µ–Ω–∏–µ –ø–æ –∑–∞–∫–∞–∑—É ‚Ññ{order_id}:\n\n{user_msg}")
        bot.send_message(message.chat.id, "‚úÖ –°–æ–æ–±—â–µ–Ω–∏–µ —É—Å–ø–µ—à–Ω–æ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ –∫–ª–∏–µ–Ω—Ç—É.", parse_mode="HTML")
    except Exception as e:
        bot.send_message(message.chat.id, f"‚ùå –ù–µ —É–¥–∞–ª–æ—Å—å –æ—Ç–ø—Ä–∞–≤–∏—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ: {e}", parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üë§ –ú–æ–π –ø—Ä–æ—Ñ–∏–ª—å")
def show_profile(message):
    """
    –ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –ª–∞–∫–æ–Ω–∏—á–Ω—É—é –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–æ—Ñ–∏–ª–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.
    –¢–µ–ø–µ—Ä—å –≤–∫–ª—é—á–∞–µ—Ç –±–∞–ª–ª—ã, –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥—Ä—É–∑–µ–π –∏ –∑–∞–∫–∞–∑–æ–≤.
    """
    telegram_id = str(message.chat.id)
    user = db_manager.get_user(telegram_id)
    if not user:
        bot.send_message(telegram_id, "‚ö†Ô∏è –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, —Å–Ω–∞—á–∞–ª–∞ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–π—Ç–µ—Å—å ‚Äî –Ω–∞–∂–º–∏—Ç–µ /start.", parse_mode="HTML")
        return
       
    points = user[5] or 0
    referrals = user[-2] or 0 # –ò–∑–º–µ–Ω–µ–Ω–æ –Ω–∞ -2, —Ç–∞–∫ –∫–∞–∫ –¥–æ–±–∞–≤–∏–ª–æ—Å—å –ø–æ–ª–µ orders_count
    orders = user[-1] or 0
   
    text = (
        f"üë§ <b>–¢–≤–æ–π –ø—Ä–æ—Ñ–∏–ª—å</b>\n\n"
        f"‚Ä¢ –ë–∞–ª–ª—ã: <b>{points}</b>\n"
        f"‚Ä¢ –ü—Ä–∏–≥–ª–∞—à–µ–Ω–æ –¥—Ä—É–∑–µ–π: <b>{referrals}</b>\n"
        f"‚Ä¢ –°–¥–µ–ª–∞–Ω–æ –∑–∞–∫–∞–∑–æ–≤: <b>{orders}</b>\n\n"
        "‚ÑπÔ∏è –ë–∞–ª–ª—ã –Ω–∞—á–∏—Å–ª—è—é—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –ø–æ—Å–ª–µ –æ–ø–ª–∞—Ç—ã –∑–∞–∫–∞–∑–∞. –ü–æ–¥—Ä–æ–±–Ω–µ–µ ‚Äî –≤ —Ä–∞–∑–¥–µ–ª–µ ¬´üîó –†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –ø—Ä–æ–≥—Ä–∞–º–º–∞¬ª.")
    bot.send_message(telegram_id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üîó –†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –ø—Ä–æ–≥—Ä–∞–º–º–∞")
def show_referral_program(message):
    """
    –ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω—É—é –ø—Ä–æ–≥—Ä–∞–º–º—É, –≤–∫–ª—é—á–∞—è –ø–æ–¥—Ä–æ–±–Ω—É—é –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —á–µ–ª–ª–µ–Ω–¥–∂–µ.
    """
    telegram_id = str(message.chat.id)
    user = db_manager.get_user(telegram_id)
    if not user:
        bot.send_message(telegram_id, "‚ö†Ô∏è –°–Ω–∞—á–∞–ª–∞ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–π—Ç–µ—Å—å ‚Äî –Ω–∞–∂–º–∏—Ç–µ /start.", parse_mode="HTML")
        return
    points = user[5] or 0
    referrals = user[-2] or 0
   
    bot_username = bot.get_me().username or "your_bot"
    link = f"https://t.me/{bot_username}?start={telegram_id}"
    text = (
        "üîó <b>–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –ø—Ä–æ–≥—Ä–∞–º–º–∞</b>\n\n"
        f"–ü—Ä–∏–≥–ª–∞—à–∞–π –¥—Ä—É–∑–µ–π –∏ –ø–æ–ª—É—á–∞–π <b>{REFERRAL_BONUS}</b> –±–∞–ª–ª–æ–≤ –∑–∞ –∫–∞–∂–¥–æ–≥–æ, –∫—Ç–æ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–µ—Ç—Å—è –ø–æ —Ç–≤–æ–µ–π —Å—Å—ã–ª–∫–µ –∏ —Å–¥–µ–ª–∞–µ—Ç –ø–µ—Ä–≤—ã–π –ø–ª–∞—Ç–Ω—ã–π –∑–∞–∫–∞–∑.\n\n"
        f"üîó –¢–≤–æ—è —Å—Å—ã–ª–∫–∞:\n{link}\n\n"
        "‚ú® <b>–ß–µ–ª–ª–µ–Ω–¥–∂ ‚Äî 15% —Å–∫–∏–¥–∫–∞</b>\n"
        "–ß—Ç–æ–±—ã –ø–æ–ª—É—á–∏—Ç—å —Ä–∞–∑–æ–≤—É—é —Å–∫–∏–¥–∫—É 15% –Ω—É–∂–Ω–æ:\n"
        "‚Ä¢ –ù–∞–∫–æ–ø–∏—Ç—å 500 –±–∞–ª–ª–æ–≤\n"
        "‚Ä¢ –ü—Ä–∏–≥–ª–∞—Å–∏—Ç—å 10 –¥—Ä—É–∑–µ–π\n\n"
        "<i>–ë–æ–Ω—É—Å—ã –Ω–∞—á–∏—Å–ª—è—é—Ç—Å—è —Ç–æ–ª—å–∫–æ –ø–æ—Å–ª–µ –ø–µ—Ä–≤–æ–π –ø–æ–∫—É–ø–∫–∏ —Ä–µ—Ñ–µ—Ä–∞–ª–∞ (—á—Ç–æ–±—ã –ø—Ä–µ–¥–æ—Ç–≤—Ä–∞—Ç–∏—Ç—å —Ñ–µ–π–∫–æ–≤—ã–µ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏).</i>"
    )
    bot.send_message(telegram_id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üõ† –¢–µ—Ö–ø–æ–¥–¥–µ—Ä–∂–∫–∞")
def support_info(message):
    """–ü—Ä–µ–¥–æ—Å—Ç–∞–≤–ª—è–µ—Ç –∫–æ–Ω—Ç–∞–∫—Ç —Ç–µ—Ö–ø–æ–¥–¥–µ—Ä–∂–∫–∏."""
    bot.send_message(
        message.chat.id,
        "üõ† –¢–µ—Ö–ø–æ–¥–¥–µ—Ä–∂–∫–∞: @tamiklung\n–ï—Å–ª–∏ –µ—Å—Ç—å –ø—Ä–æ–±–ª–µ–º—ã ‚Äî –æ–ø–∏—à–∏—Ç–µ —Å–∏—Ç—É–∞—Ü–∏—é, –º—ã –æ—Ç–≤–µ—Ç–∏–º.",
        parse_mode="HTML"
    )
@bot.message_handler(func=lambda m: m.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –µ—â—ë")
def add_more(message):
    """–ü—Ä–µ–¥–ª–∞–≥–∞–µ—Ç –¥–æ–±–∞–≤–∏—Ç—å –µ—â–µ –Ω–∞–ø–∏—Ç–∫–∏ –≤ –∫–æ—Ä–∑–∏–Ω—É."""
    bot.send_message(
        message.chat.id,
        "–û—Ç–ª–∏—á–Ω–æ! –ù–∞–ø–∏—à–∏ –Ω–∞–∑–≤–∞–Ω–∏–µ —Å–ª–µ–¥—É—é—â–µ–≥–æ –Ω–∞–ø–∏—Ç–∫–∞ (–ø—Ä–∏–º–µ—Ä: <i>–õ–∞—Ç—Ç–µ 0.3–ª</i>) –∏–ª–∏ –≤—ã–±–µ—Ä–∏ –∏–∑ –º–µ–Ω—é.",
        reply_markup=main_keyboard(),
        parse_mode="HTML"
    )
@bot.message_handler(func=lambda m: m.text == "‚úÖ –ü–µ—Ä–µ–π—Ç–∏ –∫ –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—é")
def go_to_checkout(message):
    """–ü–µ—Ä–µ–≤–æ–¥–∏—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∫ –æ—Ñ–æ—Ä–º–ª–µ–Ω–∏—é –∑–∞–∫–∞–∑–∞."""
    show_cart(message)
# --- –û–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –¥–ª—è –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏ ---
@bot.message_handler(commands=["admin"])
def admin_panel(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞", "‚ö†Ô∏è –ù–∏–∑–∫–∏–µ –æ—Å—Ç–∞—Ç–∫–∏")
    kb.add("‚öôÔ∏è –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞–º–∏")
    kb.add("üßæ –ü–æ—Å–ª–µ–¥–Ω–∏–µ –∑–∞–∫–∞–∑—ã")
    kb.add("üìã –ü—Ä–æ—Å–º–æ—Ç—Ä –º–µ–Ω—é")
    bot.send_message(message.chat.id, "üî• –ê–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å –∞–∫—Ç–∏–≤–Ω–∞. –í—ã–±–µ—Ä–∏—Ç–µ –∫–æ–º–∞–Ω–¥—É:", reply_markup=kb)
@bot.message_handler(func=lambda m: m.text == "üìã –ü—Ä–æ—Å–º–æ—Ç—Ä –º–µ–Ω—é")
def show_admin_menu(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Ç–µ–∫—É—â–µ–µ –º–µ–Ω—é –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    rows = db_manager.get_menu()
    if not rows:
        bot.send_message(message.chat.id, "üì≠ –ú–µ–Ω—é –ø–æ–∫–∞ –ø—É—Å—Ç–æ–µ.", parse_mode="HTML")
        return
    text = "üìã <b>–¢–µ–∫—É—â–µ–µ –º–µ–Ω—é:</b>\n"
    current_cat = None
    for cat, name, size, price, qty in rows:
        if cat != current_cat:
            text += f"\nüî∏ <b>{cat}</b>\n"
            current_cat = cat
        size_str = f" {size}–ª" if size else ""
        text += f"‚Ä¢ {name}{size_str} ‚Äî {price}‚ÇΩ (–æ—Å—Ç–∞–ª–æ—Å—å: {qty})\n"
    bot.send_message(message.chat.id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "‚öôÔ∏è –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞–º–∏")
def manage_items_menu(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –º–µ–Ω—é —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ç–æ–≤–∞—Ä–∞–º–∏."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("‚ûï –î–æ–±–∞–≤–∏—Ç—å –≤—Ä—É—á–Ω—É—é", callback_data="admin_add_manual"))
    kb.add(types.InlineKeyboardButton("‚¨ÜÔ∏è –ó–∞–≥—Ä—É–∑–∏—Ç—å –∏–∑ Excel", callback_data="admin_add_excel"))
    kb.add(types.InlineKeyboardButton("‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ", callback_data="admin_update_qty"))
    kb.add(types.InlineKeyboardButton("üóëÔ∏è –£–¥–∞–ª–∏—Ç—å", callback_data="admin_delete"))
    bot.send_message(message.chat.id, "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ç–æ–≤–∞—Ä–∞–º–∏:", reply_markup=kb)
@bot.callback_query_handler(func=lambda call: call.data == "admin_add_manual")
def admin_add_prompt_callback(call):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –¥–ª—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —Ç–æ–≤–∞—Ä–∞ –≤—Ä—É—á–Ω—É—é."""
    msg = bot.send_message(call.message.chat.id, "‚úèÔ∏è –í–≤–µ–¥–∏ –æ–¥–Ω—É –ø–æ–∑–∏—Ü–∏—é –≤ —Ñ–æ—Ä–º–∞—Ç–µ:\n–ö–∞—Ç–µ–≥–æ—Ä–∏—è;–ù–∞–∑–≤–∞–Ω–∏–µ;–†–∞–∑–º–µ—Ä;–¶–µ–Ω–∞;–û—Å—Ç–∞—Ç–æ–∫\n\n–ü—Ä–∏–º–µ—Ä: Classic;–õ–∞—Ç—Ç–µ;0.3;250;10", parse_mode="HTML")
    bot.register_next_step_handler(msg, apply_new_item)
def apply_new_item(message):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –Ω–æ–≤—ã–π —Ç–æ–≤–∞—Ä –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    try:
        cat, name, size, price, qty = [x.strip() for x in message.text.split(";")]
        price = int(price); qty = int(qty)
        db_manager.admin_update_item(cat, name, size, price, qty)
        bot.send_message(message.chat.id, f"‚úÖ –î–æ–±–∞–≤–ª–µ–Ω–æ: {cat} | {name} {size}–ª ‚Äî {price}‚ÇΩ (–æ—Å—Ç–∞—Ç–æ–∫ {qty})")
    except Exception as e:
        bot.send_message(message.chat.id, f"‚ùå –û—à–∏–±–∫–∞: {e}\n–ü—Ä–∏–º–µ—Ä –ø—Ä–∞–≤–∏–ª—å–Ω–æ–≥–æ —Ñ–æ—Ä–º–∞—Ç–∞: Classic;–õ–∞—Ç—Ç–µ;0.3;250;10", parse_mode="HTML")
@bot.callback_query_handler(func=lambda call: call.data == "admin_add_excel")
def admin_add_excel_prompt(call):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç Excel-—Ñ–∞–π–ª –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –º–µ–Ω—é."""
    msg = bot.send_message(call.message.chat.id, "üìÑ –ü—Ä–∏—à–ª–∏—Ç–µ Excel (.xlsx) –≤ —Ñ–æ—Ä–º–∞—Ç–µ: –ö–∞—Ç–µ–≥–æ—Ä–∏—è | –ù–∞–∑–≤–∞–Ω–∏–µ | –†–∞–∑–º–µ—Ä | –¶–µ–Ω–∞ | –û—Å—Ç–∞—Ç–æ–∫", parse_mode="HTML")
    bot.register_next_step_handler(msg, process_excel_upload)
def process_excel_upload(message):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–π Excel-—Ñ–∞–π–ª –∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç –º–µ–Ω—é."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    if not message.document or not message.document.file_name.endswith(('.xlsx', '.xls')):
        bot.send_message(message.chat.id, "‚ùå –≠—Ç–æ –Ω–µ Excel-—Ñ–∞–π–ª. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â—ë —Ä–∞–∑.", parse_mode="HTML")
        return
   
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
       
        temp_file_path = f"/tmp/{message.document.file_id}.xlsx"
        with open(temp_file_path, 'wb') as f:
            f.write(downloaded_file)
           
        workbook = openpyxl.load_workbook(temp_file_path)
        sheet = workbook.active
       
        items_added = 0
        with db_manager.get_conn() as conn:
            cur = conn.cursor()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 5:
                    cat, name, size, price, qty = row[:5]
                    try:
                        price = int(price)
                        qty = int(qty)
                        cur.execute("INSERT OR REPLACE INTO stock (category, name, size, price, quantity) VALUES (?, ?, ?, ?, ?)",
                                    (cat.strip(), name.strip(), str(size).strip(), price, qty))
                        items_added += 1
                    except (ValueError, TypeError):
                        continue
            conn.commit()
       
        os.remove(temp_file_path)
        bot.send_message(message.chat.id, f"‚úÖ –ì–æ—Ç–æ–≤–æ ‚Äî –æ–±–Ω–æ–≤–ª–µ–Ω–æ {items_added} –ø–æ–∑–∏—Ü–∏–π –∏–∑ —Ñ–∞–π–ª–∞.")
       
    except Exception as e:
        bot.send_message(message.chat.id, f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Ñ–∞–π–ª–∞: {e}", parse_mode="HTML")
@bot.callback_query_handler(func=lambda call: call.data == "admin_update_qty")
def admin_update_qty_prompt(call):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –¥–ª—è –∏–∑–º–µ–Ω–µ–Ω–∏—è –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ —Ç–æ–≤–∞—Ä–∞."""
    msg = bot.send_message(call.message.chat.id, "‚úèÔ∏è –í–≤–µ–¥–∏: –ù–∞–∑–≤–∞–Ω–∏–µ –†–∞–∑–º–µ—Ä, –ù–æ–≤–æ–µ_–∫–æ–ª–∏—á–µ—Å—Ç–≤–æ\n–ü—Ä–∏–º–µ—Ä: –õ–∞—Ç—Ç–µ 0.3, 15", parse_mode="HTML")
    bot.register_next_step_handler(msg, apply_update_qty)
def apply_update_qty(message):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–∞ –≤ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö."""
    if message.chat.id != ADMIN_GROUP_ID:
        return
    try:
        parts = [p.strip() for p in message.text.split(",")]
        if len(parts) != 2: raise ValueError
        name_size_str, new_qty_str = parts
        name_parts = name_size_str.rsplit(" ", 1)
        if len(name_parts) != 2: raise ValueError
        name, size = name_parts[0], name_parts[1].replace("–ª", "")
        new_qty = int(new_qty_str)
        if db_manager.admin_update_qty(name, size, new_qty):
            bot.send_message(message.chat.id, f"‚úÖ –û—Å—Ç–∞—Ç–æ–∫ –¥–ª—è {name} {size}–ª –æ–±–Ω–æ–≤–ª–µ–Ω –¥–æ {new_qty}.")
        else:
            bot.send_message(message.chat.id, f"‚ùå –¢–æ–≤–∞—Ä '{name} {size}–ª' –Ω–µ –Ω–∞–π–¥–µ–Ω.")
    except Exception:
        bot.send_message(message.chat.id, "‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç. –ü—Ä–∏–º–µ—Ä: –õ–∞—Ç—Ç–µ 0.3, 15", parse_mode="HTML")
@bot.callback_query_handler(func=lambda call: call.data == "admin_delete")
def admin_delete_prompt(call):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è —Ç–æ–≤–∞—Ä–∞."""
    msg = bot.send_message(call.message.chat.id, "üóë –í–≤–µ–¥–∏—Ç–µ: –ù–∞–∑–≤–∞–Ω–∏–µ –†–∞–∑–º–µ—Ä (–ø—Ä–∏–º–µ—Ä: –õ–∞—Ç—Ç–µ 0.3–ª)", parse_mode="HTML")
    bot.register_next_step_handler(msg, apply_delete_item)
def apply_delete_item(message):
    """–£–¥–∞–ª—è–µ—Ç —Ç–æ–≤–∞—Ä –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö."""
    if message.chat.id != ADMIN_GROUP_ID: return
    try:
        name_size_str = message.text.strip()
        name_parts = name_size_str.rsplit(" ", 1)
        if len(name_parts) != 2: raise ValueError
        name, size = name_parts[0], name_parts[1].replace("–ª", "")
        if db_manager.admin_delete_item(name, size):
            bot.send_message(message.chat.id, f"‚úÖ –¢–æ–≤–∞—Ä '{name} {size}–ª' —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª—ë–Ω.")
        else:
            bot.send_message(message.chat.id, f"‚ùå –¢–æ–≤–∞—Ä '{name} {size}–ª' –Ω–µ –Ω–∞–π–¥–µ–Ω.")
    except Exception:
        bot.send_message(message.chat.id, "‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â—ë —Ä–∞–∑.", parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞")
def show_admin_stats(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç –æ–±—â—É—é —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
    if message.chat.id != ADMIN_GROUP_ID: return
    order_count, total_revenue, user_count = db_manager.get_admin_data()
    text = (
        "üìä <b>–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞</b>\n\n"
        f"‚Ä¢ –û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–∫–∞–∑–æ–≤: <b>{order_count or 0}</b>\n"
        f"‚Ä¢ –û–±—â–∞—è –≤—ã—Ä—É—á–∫–∞: <b>{total_revenue or 0}‚ÇΩ</b>\n"
        f"‚Ä¢ –í—Å–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: <b>{user_count or 0}</b>"
    )
    bot.send_message(message.chat.id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "‚ö†Ô∏è –ù–∏–∑–∫–∏–µ –æ—Å—Ç–∞—Ç–∫–∏")
def show_low_stock(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Ç–æ–≤–∞—Ä—ã —Å –Ω–∏–∑–∫–∏–º –æ—Å—Ç–∞—Ç–∫–æ–º –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
    if message.chat.id != ADMIN_GROUP_ID: return
    low_stock_items = db_manager.get_low_stock()
    if not low_stock_items:
        bot.send_message(message.chat.id, "‚úÖ –ù–∞ —Å–∫–ª–∞–¥–µ –Ω–µ—Ç —Ç–æ–≤–∞—Ä–æ–≤ —Å –Ω–∏–∑–∫–∏–º –æ—Å—Ç–∞—Ç–∫–æ–º.", parse_mode="HTML")
        return
    text = "‚ö†Ô∏è <b>–¢–æ–≤–∞—Ä—ã —Å –Ω–∏–∑–∫–∏–º –æ—Å—Ç–∞—Ç–∫–æ–º (&lt;3 —à—Ç.):</b>\n"
    for cat, name, size, qty in low_stock_items:
        size_str = f" {size}–ª" if size else ""
        text += f"‚Ä¢ {name}{size_str} ‚Äî {qty} —à—Ç\n"
    bot.send_message(message.chat.id, text, parse_mode="HTML")
@bot.message_handler(func=lambda m: m.text == "üßæ –ü–æ—Å–ª–µ–¥–Ω–∏–µ –∑–∞–∫–∞–∑—ã")
def show_recent_orders(message):
    """–ü–æ–∫–∞–∑—ã–≤–∞–µ—Ç —Å–ø–∏—Å–æ–∫ –ø–æ—Å–ª–µ–¥–Ω–∏—Ö –∑–∞–∫–∞–∑–æ–≤ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏."""
    if message.chat.id != ADMIN_GROUP_ID: return
    orders = db_manager.get_recent_orders()
    if not orders:
        bot.send_message(message.chat.id, "üßæ –ü–æ–∫–∞ –Ω–µ—Ç –≤—ã–ø–æ–ª–Ω–µ–Ω–Ω—ã—Ö –∑–∞–∫–∞–∑–æ–≤.", parse_mode="HTML")
        return
    text = "üßæ <b>–ü–æ—Å–ª–µ–¥–Ω–∏–µ 10 –∑–∞–∫–∞–∑–æ–≤:</b>\n\n"
    for order_id, user_name, total, created_at in orders:
        try:
            dt_obj = datetime.datetime.strptime(created_at.split(".")[0], "%Y-%m-%d %H:%M:%S")
            nice_dt = dt_obj.strftime('%H:%M:%S %d.%m.%Y')
        except Exception:
            nice_dt = created_at
        text += (
            f"<b>‚Ññ{order_id}</b> ‚Äî <b>{user_name}</b>\n"
            f" –°—É–º–º–∞: {total}‚ÇΩ\n"
            f" –í—Ä–µ–º—è: {nice_dt}\n\n"
        )
    bot.send_message(message.chat.id, text, parse_mode="HTML")
@bot.message_handler(content_types=["text"])
def handle_text_message(message):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç —Ç–µ–∫—Å—Ç–æ–≤—ã–µ —Å–æ–æ–±—â–µ–Ω–∏—è –¥–ª—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —Ç–æ–≤–∞—Ä–æ–≤ –≤ –∫–æ—Ä–∑–∏–Ω—É."""
    telegram_id = str(message.chat.id)
    parts = message.text.strip().rsplit(" ", 1)
    if len(parts) < 2 or not parts[1].endswith("–ª"):
        # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –Ω–µ—Ñ–æ—Ä–º–∞—Ç–Ω—ã–µ —Å–æ–æ–±—â–µ–Ω–∏—è ‚Äî –±–æ—Ç –æ—Ä–∏–µ–Ω—Ç–∏—Ä–æ–≤–∞–Ω –Ω–∞ –∫–æ–º–∞–Ω–¥—ã/–∫–Ω–æ–ø–∫–∏
        return
    name = parts[0]
    size_str = parts[1].replace("–ª", "")
   
    price_qty = db_manager.get_stock_by_name_size(name, size_str)
    if not price_qty:
        bot.send_message(telegram_id, f"‚ùå –¢–æ–≤–∞—Ä '{message.text}' –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü—Ä–æ–≤–µ—Ä—å —Ä–µ–≥–∏—Å—Ç—Ä/–Ω–∞–∑–≤–∞–Ω–∏–µ.", parse_mode="HTML")
        return
    price, qty = price_qty
    if price is None or qty <= 0:
        bot.send_message(telegram_id, f"‚ùå –ò–∑–≤–∏–Ω–∏—Ç–µ, —Ç–æ–≤–∞—Ä–∞ '{message.text}' –Ω–µ—Ç –≤ –Ω–∞–ª–∏—á–∏–∏.", parse_mode="HTML")
        return
    db_manager.add_to_cart(telegram_id, f"{name} {size_str}–ª", price, 1)
    bot.send_message(telegram_id, f"‚úÖ –î–æ–±–∞–≤–∏–ª <b>{name} {size_str}–ª</b> –≤ –∫–æ—Ä–∑–∏–Ω—É. –•–æ—Ç–∏—Ç–µ –¥–æ–±–∞–≤–∏—Ç—å –µ—â—ë?", reply_markup=handle_add_more_keyboard(), parse_mode="HTML")
def main():
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(f"‚ùå An error occurred: {e}")
        time.sleep(5)
        main()
if __name__ == '__main__':
    main()
